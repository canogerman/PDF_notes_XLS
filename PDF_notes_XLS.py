import os
import re
from openpyxl import Workbook
import fitz  # PyMuPDF
import locale

# Configura el idioma predeterminado para el análisis de fechas
locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

# Crea una instancia de un libro de Excel y obtiene la hoja activa
wb = Workbook()
ws = wb.active

# Crea las cabeceras de las columnas en la hoja de Excel
ws["A1"] = "Número de Nota"
ws["B1"] = "Referencia"
ws["C1"] = "Fecha"
ws["D1"] = "Texto"

# Busca todos los archivos PDF en una carpeta específica
pdf_folder = r"C:\Notas"
pdf_files = [f for f in os.listdir(pdf_folder) if f.endswith(".pdf")]

# Expresión regular específica para el formato "Miércoles 13 de Diciembre de 2023"
date_regex = re.compile(r"(Domingo|Lunes|Martes|Miércoles|Jueves|Viernes|Sábado) (\d{1,2}) de (Enero|Febrero|Marzo|Abril|Mayo|Junio|Julio|Agosto|Septiembre|Octubre|Noviembre|Diciembre) de (\d{4})")

# Diccionario para mapear nombres de meses en español a números de mes
meses = {
    'Enero': '01',
    'Febrero': '02',
    'Marzo': '03',
    'Abril': '04',
    'Mayo': '05',
    'Junio': '06',
    'Julio': '07',
    'Agosto': '08',
    'Septiembre': '09',
    'Octubre': '10',
    'Noviembre': '11',
    'Diciembre': '12'
}

# Lee cada archivo PDF y extrae la información necesaria
row = 2  # Empieza en la segunda fila de la hoja de Excel
for pdf_file in pdf_files:
    pdf_path = os.path.join(pdf_folder, pdf_file)

    # Utiliza el bloque "with" para abrir el archivo PDF
    with fitz.open(pdf_path) as pdf_document:

        # Inicializa las variables para asegurarse de que estén definidas
        numero_nota = ""
        referencia = ""
        fecha_legible = ""
        texto_acumulado = ""  # Variable para acumular el texto de todas las páginas

        # Lee el contenido del PDF y busca la información requerida
        for page_num in range(pdf_document.page_count):
            page = pdf_document[page_num]
            text = page.get_text("text")

            # Acumula el texto de todas las páginas
            texto_acumulado += text
            
        # Busca los datos de interés en el texto acumulado
        if "De mi mayor consideración:" in texto_acumulado:
            # Extrae el número de nota del nombre del archivo sin la extensión
            numero_nota = os.path.splitext(pdf_file)[0]
            referencia = texto_acumulado.split("Referencia: ")[1].split("\n")[0]

            # Busca el texto comprendido entre las cadenas deseadas
            start_idx = texto_acumulado.find("De mi mayor consideración:") + len("De mi mayor consideración:")
            end_idx = texto_acumulado.find("Sin otro particular saluda atte.")
            texto = texto_acumulado[start_idx:end_idx].strip()
            
            # Realiza reemplazos para mejorar la legibilidad en Excel
            texto = texto.replace('.\n', '.\n---').replace('\n', '').replace('•', '\n• ').replace('---', '\n')

            # Busca fechas en el texto utilizando la expresión regular
            match = date_regex.search(texto_acumulado)
            if match:
                day = match.group(2)
                month = meses[match.group(3)]
                year = match.group(4)

                try:
                    # Construye la fecha en el formato correcto
                    fecha_legible = f"{year}-{month}-{day}"
                except ValueError as e:
                    print(f"Error al construir la fecha: {e}")

        # Escribe los datos en la hoja de Excel
        ws["A{}".format(row)] = numero_nota
        ws["B{}".format(row)] = referencia
        ws["C{}".format(row)] = fecha_legible
        ws["D{}".format(row)] = texto
        row += 1

# Guarda el libro de Excel en la ruta especificada
wb.save(r"C:\Notas\Notas.xlsx")

# Información del autor
autor = "Script creado por Germán Cano Novotny."

# Mensaje de información al usuario
print("Proceso completado. El archivo 'Notas.xlsx' ha sido generado. {}".format(autor))

# Pausa para que el usuario pueda leer el mensaje antes de cerrar la consola
input("Presiona Enter para salir...")